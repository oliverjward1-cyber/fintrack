#!/usr/bin/env python3
"""
Pull establishment data from the UK Food Hygiene Rating Scheme (FHRS) API
and export sales-prospect data to Excel.

For each named local authority, fetches all establishments, filters to the
target business types and hygiene ratings, and writes one spreadsheet with
the three inspection sub-scores. Note that FHRS sub-scores count points
DEDUCTED, so higher = worse: a ConfidenceInManagement of 20-30 alongside an
overall rating of 3-4 marks a business with decent practice but weak
documentation.

Rows are sorted by rating, then inspection date oldest-first (oldest are
due a revisit soonest).

Usage:
    python fhrs_prospects.py
    python fhrs_prospects.py --authorities "Plymouth" "East Devon" --ratings 2 3 4
    python fhrs_prospects.py --output devon_prospects.xlsx

Requires: requests, openpyxl   (pip install requests openpyxl)
No API key needed.
"""

import argparse
import sys
from datetime import datetime

import requests
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

API_BASE = "https://api.ratings.food.gov.uk"
HEADERS = {
    "x-api-version": "2",
    "accept": "application/json",
}
PAGE_SIZE = 5000  # API maximum

DEFAULT_AUTHORITIES = ["South Hams", "Torbay", "Exeter", "Teignbridge"]
DEFAULT_BUSINESS_TYPES = [
    "Restaurant/Cafe/Canteen",
    "Pub/bar/nightclub",
    "Takeaway/sandwich shop",
]
DEFAULT_RATINGS = ["3", "4"]

COLUMNS = [
    "Business name",
    "Address",
    "Postcode",
    "Rating",
    "Last inspection",
    "Hygiene score",
    "Structural score",
    "Confidence in management score",
]


def get_json(session, path, params=None):
    resp = session.get(f"{API_BASE}{path}", params=params, timeout=60)
    resp.raise_for_status()
    return resp.json()


def lookup_authority_ids(session, names):
    """Map authority names to LocalAuthorityIds via the /Authorities endpoint."""
    data = get_json(session, "/Authorities/basic")
    authorities = data["authorities"]
    resolved = {}
    for name in names:
        matches = [
            a for a in authorities if name.lower() in a["Name"].lower()
        ]
        if not matches:
            print(f"WARNING: no authority matching '{name}' — skipping.", file=sys.stderr)
            continue
        if len(matches) > 1:
            print(
                f"WARNING: '{name}' matched {len(matches)} authorities; "
                f"using '{matches[0]['Name']}'. Others: "
                + ", ".join(a["Name"] for a in matches[1:]),
                file=sys.stderr,
            )
        resolved[matches[0]["Name"]] = matches[0]["LocalAuthorityId"]
    return resolved


def fetch_establishments(session, authority_id):
    """Fetch every establishment for one authority, following pagination."""
    establishments = []
    page = 1
    while True:
        data = get_json(
            session,
            "/Establishments",
            params={
                "localAuthorityId": authority_id,
                "pageNumber": page,
                "pageSize": PAGE_SIZE,
            },
        )
        establishments.extend(data["establishments"])
        meta = data.get("meta", {})
        if page >= meta.get("totalPages", 1):
            break
        page += 1
    return establishments


def parse_rating_date(value):
    """RatingDate arrives as e.g. '2023-05-12T00:00:00', or null/'' if unknown."""
    if not value:
        return None
    try:
        return datetime.fromisoformat(value).date()
    except ValueError:
        return None


def build_rows(establishments, authority_name, business_types, ratings):
    wanted_types = {t.lower() for t in business_types}
    wanted_ratings = set(ratings)
    rows = []
    for e in establishments:
        if (e.get("BusinessType") or "").lower() not in wanted_types:
            continue
        if e.get("RatingValue") not in wanted_ratings:
            continue
        address = ", ".join(
            part.strip()
            for part in (
                e.get("AddressLine1"),
                e.get("AddressLine2"),
                e.get("AddressLine3"),
                e.get("AddressLine4"),
            )
            if part and part.strip()
        )
        scores = e.get("scores") or {}
        rows.append(
            {
                "authority": authority_name,
                "name": e.get("BusinessName", ""),
                "address": address,
                "postcode": e.get("PostCode", "") or "",
                "rating": int(e["RatingValue"]),
                "inspected": parse_rating_date(e.get("RatingDate")),
                "hygiene": scores.get("Hygiene"),
                "structural": scores.get("Structural"),
                "management": scores.get("ConfidenceInManagement"),
            }
        )
    return rows


def write_excel(rows, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Prospects"

    header = ["Local authority"] + COLUMNS
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for r in rows:
        ws.append(
            [
                r["authority"],
                r["name"],
                r["address"],
                r["postcode"],
                r["rating"],
                r["inspected"],
                r["hygiene"],
                r["structural"],
                r["management"],
            ]
        )

    # Date display format for the inspection column (column F)
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
        row[0].number_format = "DD/MM/YYYY"

    widths = [18, 38, 45, 10, 8, 14, 13, 15, 28]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description="Export FHRS establishment data for given local authorities to Excel."
    )
    parser.add_argument(
        "--authorities",
        nargs="+",
        default=DEFAULT_AUTHORITIES,
        help="Local authority names (substring match, case-insensitive). "
        f"Default: {', '.join(DEFAULT_AUTHORITIES)}",
    )
    parser.add_argument(
        "--business-types",
        nargs="+",
        default=DEFAULT_BUSINESS_TYPES,
        help="Business types to include (exact name, case-insensitive). "
        f"Default: {', '.join(DEFAULT_BUSINESS_TYPES)}",
    )
    parser.add_argument(
        "--ratings",
        nargs="+",
        default=DEFAULT_RATINGS,
        help=f"Hygiene ratings to include. Default: {' '.join(DEFAULT_RATINGS)}",
    )
    parser.add_argument(
        "--output",
        default=f"fhrs_prospects_{datetime.now():%Y-%m-%d}.xlsx",
        help="Output .xlsx path. Default: fhrs_prospects_<today>.xlsx",
    )
    args = parser.parse_args()

    session = requests.Session()
    session.headers.update(HEADERS)

    print("Looking up authority IDs...")
    authority_ids = lookup_authority_ids(session, args.authorities)
    if not authority_ids:
        sys.exit("No authorities resolved — nothing to do.")

    all_rows = []
    for name, auth_id in authority_ids.items():
        print(f"Fetching establishments for {name} (id {auth_id})...")
        establishments = fetch_establishments(session, auth_id)
        rows = build_rows(establishments, name, args.business_types, args.ratings)
        print(f"  {len(establishments)} establishments, {len(rows)} match filters")
        all_rows.extend(rows)

    # Rating ascending, then oldest inspection first; never-inspected rows last
    all_rows.sort(
        key=lambda r: (r["rating"], r["inspected"] is None, r["inspected"] or datetime.max.date())
    )

    write_excel(all_rows, args.output)
    print(f"Wrote {len(all_rows)} rows to {args.output}")


if __name__ == "__main__":
    main()
